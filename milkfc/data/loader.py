"""DHI 檔案載入"""
import pandas as pd
import openpyxl
import logging
from pathlib import Path
from .. import config

log = logging.getLogger("milkfc.loader")

def _detect_sheet_name(fp):
    """自動偵測表單名稱（資料合約變更會在這裡優先發現）"""
    wb = openpyxl.load_workbook(fp, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    # 優先：vw_dhireport (新版)、年份字串 (舊版)、第一個表單
    for cand in ["vw_dhireport"]:
        if cand in sheets:
            return cand
    # 找年份格式
    for s in sheets:
        if s.isdigit() and 2000 <= int(s) <= 2100:
            return s
    return sheets[0]

def _load_one_xlsx(fp: Path) -> pd.DataFrame:
    sheet = _detect_sheet_name(fp)
    log.info(f"Loading {fp.name} (sheet={sheet})")
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = rows[0]
    df = pd.DataFrame(rows[1:], columns=header)
    df = df.rename(columns=config.COLUMN_MAP)
    keep = [c for c in config.COLUMN_MAP.values() if c in df.columns]
    return df[keep]

def _normalize_types(df: pd.DataFrame) -> pd.DataFrame:
    date_cols = ["birth_date","last_calving_date","sample_date","test_date",
                 "last_breeding_date","prev_calving_date","first_breeding_date"]
    for c in date_cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")
    num_cols = ["parity","dim","milk_kg","fat_pct","protein_pct","lactose_pct",
                "scc","milk_305","age_month","breeding_count","year","month"]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    df["farm_id"] = df["farm_id"].astype(str)
    df["cow_id"] = df["cow_id"].astype(str)
    return df

def load_dhi_files(paths: list) -> pd.DataFrame:
    """載入多份 DHI xlsx，回傳合併、型別正規化後的 dataframe。"""
    frames = [_load_one_xlsx(Path(p)) for p in paths]
    df = pd.concat(frames, ignore_index=True)
    df = _normalize_types(df)
    df = df.sort_values(["farm_id","cow_id","sample_date"]).reset_index(drop=True)
    return df

def _find_dhi_files(raw_dir: Path) -> list:
    """只挑 DHI 年度檔案（YYYY+dhi.xlsx 或 dhi+YYYY.xlsx），
    排除 Farm.xlsx、官方季報、暫存檔。"""
    import re
    candidates = sorted(raw_dir.glob("*.xlsx"))
    dhi = []
    for p in candidates:
        name = p.name.lower()
        # 排除 Farm.xlsx、官方季報、Excel 暫存（~$ 開頭）
        if name.startswith("~$"):
            continue
        if name.startswith("farm"):
            continue
        if "在養" in p.name or "畜禽" in p.name:
            continue
        # DHI 命名規則：年份+dhi 或 dhi+年份
        if re.search(r"(\d{4})\s*dhi", name) or re.search(r"dhi\s*(\d{4})", name):
            dhi.append(p)
    return dhi


import json

def _fingerprint(paths: list) -> dict:
    """計算來源檔案集的指紋：檔名+大小+mtime。"""
    return {p.name: {"size": p.stat().st_size,
                     "mtime": p.stat().st_mtime}
            for p in paths}


def load_combined(cache_path: Path = None,
                   force_reload: bool = False) -> pd.DataFrame:
    """載入所有 raw_data 下的 DHI 年度 xlsx，可選 pickle 快取。

    Cache 自動失效：用「檔案指紋」（檔名+大小+mtime）比對，
    若清單或任何檔指紋變動就重新載入。比 mtime 比較更可靠。
    """
    paths = _find_dhi_files(config.RAW_DIR)
    if not paths:
        raise FileNotFoundError(f"No DHI xlsx in {config.RAW_DIR}")

    # 快取自動失效檢查（用指紋 not mtime）
    meta_path = Path(str(cache_path) + ".meta.json") if cache_path else None
    if cache_path and Path(cache_path).exists() and meta_path and meta_path.exists() \
            and not force_reload:
        try:
            with open(meta_path) as f:
                cached_fp = json.load(f)
            current_fp = _fingerprint(paths)
            if cached_fp == current_fp:
                log.info(f"Loading from cache: {cache_path}")
                return pd.read_pickle(cache_path)
            else:
                # 比對哪邊不同
                added = set(current_fp) - set(cached_fp)
                removed = set(cached_fp) - set(current_fp)
                changed = [k for k in (set(cached_fp) & set(current_fp))
                           if cached_fp[k] != current_fp[k]]
                if added: log.info(f"Cache stale - 新增檔案: {added}")
                if removed: log.info(f"Cache stale - 移除檔案: {removed}")
                if changed: log.info(f"Cache stale - 修改過的檔案: {changed}")
        except Exception as e:
            log.warning(f"讀 cache meta 失敗：{e}，將重建 cache")

    log.info(f"Loading {len(paths)} DHI files: {[p.name for p in paths]}")
    df = load_dhi_files(paths)
    if cache_path:
        df.to_pickle(cache_path)
        if meta_path:
            with open(meta_path, "w") as f:
                json.dump(_fingerprint(paths), f, indent=2)
        log.info(f"Cached to {cache_path}")
    return df
