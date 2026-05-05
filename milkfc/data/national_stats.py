"""官方畜禽統計季報解析 - 取得全國/各縣市真實乳牛場數與頭數。

資料來源：農業部農業統計資料查詢 https://agrstat.moa.gov.tw/
檔案命名規則：表X  YYYQN在養XX.xlsx (季報)
放置位置：raw_data/ 下，自動偵測最新一份。
"""
import pandas as pd
import openpyxl
import re
from pathlib import Path
from datetime import datetime
from .. import config

# 縣市名 → macro_region 對應
COUNTY_TO_MACRO = {
    "新北市":"北","臺北市":"北","台北市":"北","基隆市":"北",
    "桃園市":"北","新竹縣":"北","新竹市":"北","宜蘭縣":"北",
    "苗栗縣":"中","臺中市":"中","台中市":"中","彰化縣":"中","南投縣":"中",
    "雲林縣":"中",
    "嘉義縣":"南","嘉義市":"南","臺南市":"南","台南市":"南","高雄市":"南",
    "屏東縣":"南",
    "花蓮縣":"東","臺東縣":"東","台東縣":"東",
    "澎湖縣":"離島","金門縣":"離島","連江縣":"離島",
}


def find_latest_national_stats(raw_dir: Path = None) -> Path:
    """從 raw_data/ 找最新一份「按品項」官方季報 xlsx。"""
    raw_dir = raw_dir or config.RAW_DIR
    candidates = list(raw_dir.glob("*在養按品項*.xlsx"))
    if not candidates:
        return None
    def year_q(fp):
        m = re.search(r"(\d{3})Q(\d)", fp.name)
        if m:
            return (int(m.group(1)), int(m.group(2)))
        return (0, 0)
    return sorted(candidates, key=year_q, reverse=True)[0]


def find_all_national_stats(raw_dir: Path = None) -> list:
    """找出所有官方季報檔，按年季排序。"""
    raw_dir = raw_dir or config.RAW_DIR
    candidates = list(raw_dir.glob("*在養按品項*.xlsx"))
    def year_q(fp):
        m = re.search(r"(\d{3})Q(\d)", fp.name)
        if m:
            return (int(m.group(1)), int(m.group(2)))
        return (0, 0)
    return sorted(candidates, key=year_q)


def parse_all_quarterly(raw_dir: Path = None) -> pd.DataFrame:
    """解析所有可得季報，回傳合併的時間序列。

    Returns DataFrame:
        period (Timestamp 該季季末), n_farms, n_heads_total,
        n_milking_cows, n_unborn_heifers, n_breeding_bulls, source_file
    """
    files = find_all_national_stats(raw_dir)
    rows = []
    for fp in files:
        m = re.search(r"(\d{3})Q(\d)", fp.name)
        if not m:
            continue
        roc_year = int(m.group(1))
        quarter = int(m.group(2))
        # 該季季末（Q1=3/31, Q2=6/30, Q3=9/30, Q4=12/31）
        eom_month = quarter * 3
        eom_day = {3:31, 6:30, 9:30, 12:31}[eom_month]
        period = pd.Timestamp(f"{roc_year + 1911}-{eom_month:02d}-{eom_day}")

        s = get_national_summary(fp)
        if s:
            s["period"] = period
            s["roc_period"] = m.group(0)
            rows.append(s)
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows).sort_values("period").reset_index(drop=True)
    return df


def parse_national_stats(fp: Path = None) -> pd.DataFrame:
    """解析官方季報 xlsx 的「乳牛」分頁。

    Returns:
        DataFrame columns:
          county, county_en, n_farms, pct_farms, n_heads_total,
          pct_heads, n_milking_cows, n_unborn_heifers, n_breeding_bulls,
          macro_region, period (YYY-Qn), source_file
    """
    fp = fp or find_latest_national_stats()
    if fp is None or not Path(fp).exists():
        return pd.DataFrame()

    # 從檔名抽出年季
    m = re.search(r"(\d{3})Q(\d)", fp.name)
    period = f"{m.group(1)}-Q{m.group(2)}" if m else "unknown"

    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    if "乳牛" not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()
    ws = wb["乳牛"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # 從 row 5 (index 5) 開始是各縣市資料；row 5 是「臺閩地區」、row 6 「臺灣地區」
    # 其餘為各縣市
    records = []
    for row in rows[5:]:
        county = row[0]
        county_en = row[1]
        if not county:
            continue
        n_farms = _to_int(row[2])
        pct_farms = _to_float(row[3])
        n_heads_total = _to_int(row[4])
        pct_heads = _to_float(row[5])
        n_milking = _to_int(row[6])
        n_unborn = _to_int(row[7])
        n_breeding = _to_int(row[8])
        records.append({
            "county": county,
            "county_en": county_en,
            "n_farms": n_farms,
            "pct_farms": pct_farms,
            "n_heads_total": n_heads_total,
            "pct_heads": pct_heads,
            "n_milking_cows": n_milking,
            "n_unborn_heifers": n_unborn,
            "n_breeding_bulls": n_breeding,
            "macro_region": COUNTY_TO_MACRO.get(county, "其他"),
            "period": period,
            "source_file": fp.name,
        })
    return pd.DataFrame(records)


def _to_int(v):
    if v is None or v == "-": return 0
    if isinstance(v, (int, float)): return int(v)
    try:
        return int(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return 0


def _to_float(v):
    if v is None or v == "-": return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def get_national_summary(fp: Path = None) -> dict:
    """回傳「全國 / 臺灣地區」匯總數字。"""
    df = parse_national_stats(fp)
    if df.empty:
        return {}
    # 第一列是「臺閩地區」，第二列是「臺灣地區」
    taiwan = df[df["county"].isin(["臺灣地區","臺閩地區"])].iloc[0] \
             if (df["county"].isin(["臺灣地區","臺閩地區"])).any() else None
    if taiwan is None:
        # 沒有匯總列就把所有縣市相加
        excl_summary = df[~df["county"].isin(["臺閩地區","臺灣地區"])]
        return {
            "n_farms": int(excl_summary["n_farms"].sum()),
            "n_heads_total": int(excl_summary["n_heads_total"].sum()),
            "n_milking_cows": int(excl_summary["n_milking_cows"].sum()),
            "period": df["period"].iloc[0],
        }
    return {
        "n_farms": int(taiwan["n_farms"]),
        "n_heads_total": int(taiwan["n_heads_total"]),
        "n_milking_cows": int(taiwan["n_milking_cows"]),
        "n_unborn_heifers": int(taiwan["n_unborn_heifers"]),
        "n_breeding_bulls": int(taiwan["n_breeding_bulls"]),
        "period": taiwan["period"],
        "source_file": taiwan["source_file"],
    }


def get_county_stats(fp: Path = None, exclude_summary: bool = True) -> pd.DataFrame:
    """回傳純各縣市的乳牛統計（去掉臺閩/臺灣匯總列）。"""
    df = parse_national_stats(fp)
    if df.empty:
        return df
    if exclude_summary:
        df = df[~df["county"].isin(["臺閩地區","臺灣地區"])].copy()
    return df.reset_index(drop=True)
